from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from django.db import transaction
from django.utils import timezone

from openpyxl import load_workbook

from core.models import (
    AsignacionEquipo,
    CargoFuncionario,
    CondicionEquipo,
    Equipo,
    EquipoMonitor,
    EstadoEquipo,
    Funcionario,
    ImportacionInventarioLote,
    ImportacionInventarioRegistro,
    Marca,
    MarcaMonitorCatalogo,
    Procesador,
    PulgadaMonitorCatalogo,
    Ram,
    SistemaOperativo,
    TamanoDiscoCatalogo,
    TipoDisco,
    TipoEquipo,
    Ubicacion,
    UnidadFuncionario,
)


CODIGOS_SHEET = "CODIGOS"
DEFAULT_EXCEL_PATH = "inventariop.xlsx"
IMPORT_SHEET_GROUPS = [
    ("EQUIPOS ACTIVOS DV", ["EQUIPOS ACTIVOS DV"]),
    ("EQUIPOS NO ACTIVOS DV", ["EQUIPOS NO ACTIVOS DV", "NO ACTIVOS"]),
    ("ARRIENDO", ["ARRIENDO", "EQUIPOS ACTIVOS ARRIENDO"]),
    ("BAJA", ["BAJA"]),
]

FIXED_SEDES = ("Ánimas", "Yungay", "La Unión", "Bouchéff")
SEDE_ALIASES = {
    "Ánimas": ("ANIMAS", "ÁNIMAS"),
    "Yungay": ("YUNGAY",),
    "La Unión": ("LA UNION", "LA UNIÓN"),
    "Bouchéff": ("BOUCHEFF", "BOUCHÉFF", "BOUCHEF", "BOUCHÉF"),
}

CODIGOS_COLUMNS = {
    "condicion_equipo_dict": "CONDICIÓN EQUIPO 1 / 2",
    "tipo_equipo_dict": "TIPO EQUIPO  1/2/OTRO",
    "marca_dict": "MARCA",
    "ram_dict": "RAM",
    "procesador_dict": "PROCESADOR",
    "so_dict": "S.O",
    "tamano_disco_dict": "TAMAÑO DISCO",
    "tipo_disco_dict": "TIPO DISCO   1/2",
    "estado_equipo_dict": "ESTADO/EQUIPO (PENDIENTE)",
    "marca_monitor_dict": "MARCA MONITOR",
    "pulgada_monitor_dict": "PULGADA MONITOR",
}

FALLBACK_CODIGOS = {
    "condicion_equipo_dict": {
        1: "SERVICIO",
        2: "ARRIENDO",
    },
    "estado_equipo_dict": {
        1: "ASIGNADO",
        2: "ROBO",
        3: "RETIRAR",
        4: "PRESTAMO",
        5: "BODEGA",
        6: "EXTRAVIADO",
    },
    "tipo_equipo_dict": {
        1: "PC",
        2: "NOTEBOOK",
        3: "PLOTTER",
        4: "DATA SHOW",
        5: "IMPRESORA",
        6: "SCANNER",
        7: "ALL IN ONE",
        8: "WEBCAM",
    },
    "marca_dict": {
        1: "OLIDATA",
        2: "HP",
        3: "DELL",
        4: "LENOVO",
        5: "EPSON",
        6: "COMPAQ",
        7: "PACKARD BELL",
        8: "SONY",
        9: "ASUS",
        10: "VIEWSONIC",
        11: "LG",
        12: "SAMSUNG",
        13: "XEROX",
        14: "ACER",
        15: "APPLE",
        16: "CISCO",
    },
    "ram_dict": {
        0: "N/A",
        1: "1GB",
        2: "2GB",
        3: "4GB",
        4: "8GB",
        5: "3GB",
        6: "6GB",
        7: "12GB",
        8: "14GB",
        9: "16GB",
        10: "32GB",
        11: "256MEGA",
        12: "512 MEGA",
    },
    "procesador_dict": {
        0: "N/A",
        1: "I3",
        2: "I5",
        3: "I7",
        4: "CORE 2 DUO",
        5: "PENTIUM IV",
        6: "RYZEN 3",
        7: "RYZEN 5",
        8: "RYZEN 7",
        9: "AMD",
        10: "ADM ATHLON",
    },
    "so_dict": {
        0: "N/A",
        1: "WIN10",
        2: "WIN7",
        3: "WIN11",
        4: "WINDXP",
    },
    "tamano_disco_dict": {
        0: "N/A",
        1: "160GB",
        2: "320GB",
        3: "480GB",
        4: "500GB",
        5: "750GB",
        6: "1TB",
        7: "2TB",
        8: "80GB",
        9: "120GB",
        10: "250GB",
        11: "256GB",
        12: "240GB",
        13: "512GB",
    },
    "tipo_disco_dict": {
        0: "N/A",
        1: "SSD",
        2: "HDD",
        3: "M.2",
    },
    "marca_monitor_dict": {
        1: "OLIDATA",
        2: "HP",
        3: "DELL",
        4: "LENOVO",
        5: "EPSON",
        6: "COMPAQ",
        7: "PACKARD BELL",
        8: "SONY",
        9: "ASUS",
        10: "VIEWSONIC",
        11: "LG",
        12: "SAMSUNG",
        13: "XEROX",
        14: "ACER",
        15: "APPLE",
        16: "CISCO",
    },
    "pulgada_monitor_dict": {
        0: "N/A",
        1: "15'",
        2: "17'",
        3: "19'",
        4: "20'",
        5: "21'",
        6: "22'",
        7: "24'",
        8: "10'",
        9: "16'",
        10: "16'",
        11: "28'",
        12: "32'",
        13: "43'",
        14: "27'",
    },
}

EQUIPOS_COLUMNS = {
    "numero_serie": "SERIE ( E)",
    "numero_inventario": "SIGAC ( E)",
    "condicion_codigo": "CONDICIÓN EQUIPO 1 / 2",
    "tipo_equipo_codigo": "TIPO EQUIPO  1/2/OTRO",
    "marca_codigo": "MARCA",
    "ram_codigo": "RAM",
    "procesador_codigo": "PROCESADOR",
    "so_codigo": "S.O",
    "tamano_disco_codigo": "TAMAÑO DISCO",
    "tipo_disco_codigo": "TIPO DISCO   1/2",
    "estado_actual": "ESTADO ACTUAL   ASIGNADO/ BODEGA /EXTRAVIADO/ROBO",
    "nombre_equipo": "NOMBRE DE EQUIPO",
    "ip_maquina": "IP DE MAQUINA",
    "direccion_oficina_piso": "DIRECCIÓN OFICINA / PISO",
    "rut": "RUT",
    "nombre": "NOMBRE",
    "unidad": "DEPTO / UNIDAD",
    "cargo": "CARGO FUNCIONAL",
    "tipo_contrato": "TIPO DE CONTRATO",
    "fecha_asignacion_actualizacion": "FECHA DE ASIGNACIÓN / ACTUALIZACIÓN",
    "nombre_responsable_asignacion": "NOMBRE RESPONSABLE DE ASIGNACIÓN",
    "numero_acta": "NUMERO DE ACTA",
    "fecha_entrega": "FECHA DE ENTREGA",
    "monitor_1_marca": "MARCA MONITOR 1",
    "monitor_1_serie": "SERIE MONITOR 1",
    "monitor_1_sigac": "SIGAC MONITOR 1",
    "monitor_1_pulgada": "PULGADA MONITOR 1",
    "monitor_2_marca": "MARCA MONITOR 2",
    "monitor_2_serie": "SERIE MONITOR 2",
    "monitor_2_sigac": "SIGAC MONITOR 2",
    "monitor_2_pulgada": "PULGADA MONITOR 2",
    "observacion": "OBSERVACIÓN",
}


@dataclass
class ImportSummary:
    total_rows: int = 0
    funcionarios_created: int = 0
    funcionarios_updated: int = 0
    equipos_created: int = 0
    equipos_updated: int = 0
    asignaciones_created: int = 0
    monitores_created: int = 0
    skipped: int = 0
    created: int = 0
    updated: int = 0
    errores: list[str] = field(default_factory=list)
    skipped_rows: list[dict] = field(default_factory=list)

    def as_dict(self):
        return {
            "total_rows": self.total_rows,
            "funcionarios_created": self.funcionarios_created,
            "funcionarios_updated": self.funcionarios_updated,
            "equipos_created": self.equipos_created,
            "equipos_updated": self.equipos_updated,
            "asignaciones_created": self.asignaciones_created,
            "monitores_created": self.monitores_created,
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "errores": self.errores,
            "skipped_rows": self.skipped_rows,
        }


class ExcelInventoryImporter:
    TRACKED_MODELS = {
        "Funcionario": Funcionario,
        "Equipo": Equipo,
        "EquipoMonitor": EquipoMonitor,
        "AsignacionEquipo": AsignacionEquipo,
    }

    def __init__(self, path: str | Path | None = None):
        self.path = self._resolve_path(path or DEFAULT_EXCEL_PATH)
        self.lote = None

    def run(self):
        wb = load_workbook(self.path, data_only=True)
        codigos_sheet = self._get_sheet(wb, CODIGOS_SHEET)
        self.lote = ImportacionInventarioLote.objects.create(
            archivo_origen=self.path.name,
            estado="procesando",
        )

        dictionaries = self._build_codigos_dicts(codigos_sheet)
        summary = ImportSummary()

        for _, candidates in IMPORT_SHEET_GROUPS:
            sheet = self._find_first_sheet(wb, candidates)
            if sheet is None:
                continue
            self._import_equipos(sheet, dictionaries, summary)

        self.lote.estado = "completado"
        self.lote.resumen = summary.as_dict()
        with transaction.atomic():
            self.lote.save(update_fields=["estado", "resumen"])
        return summary.as_dict()

    def _resolve_path(self, path):
        candidate = Path(path).expanduser()
        if not candidate.is_absolute():
            candidate = Path.cwd() / candidate
        return candidate

    def _get_sheet(self, workbook, name):
        if name not in workbook.sheetnames:
            raise ValueError(f'La hoja "{name}" no existe en "{self.path.name}".')
        return workbook[name]

    def _find_first_sheet(self, workbook, candidates):
        for candidate in candidates:
            if candidate in workbook.sheetnames:
                return workbook[candidate]
        return None

    def _build_codigos_dicts(self, sheet):
        header = self._header_map(sheet)
        dicts = {key: values.copy() for key, values in FALLBACK_CODIGOS.items()}
        for key in CODIGOS_COLUMNS:
            dicts.setdefault(key, {})

        for row in sheet.iter_rows(min_row=2, values_only=True):
            for dict_name, column_name in CODIGOS_COLUMNS.items():
                col_idx = header.get(self._normalize_header(column_name))
                if col_idx is None or col_idx >= len(row):
                    continue
                codigo, descripcion = self._parse_code_and_description(row[col_idx])
                if codigo is None or not descripcion:
                    continue
                dicts[dict_name][codigo] = descripcion

        return dicts

    def _import_equipos(self, sheet, dictionaries, summary):
        header = self._header_map(sheet)
        original_headers = self._original_headers(sheet)

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = self._row_to_dict(row, original_headers, sheet.title, row_number, dictionaries)

            if self._row_is_empty(row):
                summary.skipped += 1
                self._skip_row(summary, sheet.title, row_number, "Fila vacía.")
                continue

            summary.total_rows += 1

            try:
                # Each row runs in its own savepoint so one DB failure does not
                # poison the whole import transaction.
                with transaction.atomic():
                    funcionario, funcionario_action = self._upsert_funcionario(row, header, summary, row_data)
                    equipo, equipo_action = self._upsert_equipo(row, header, dictionaries, summary, row_data, sheet.title, row_number)
                    if equipo is None:
                        summary.skipped += 1
                        self._skip_row(summary, sheet.title, row_number, "No fue posible identificar equipo (sin SIGAC ni serie).")
                        continue
                    monitores_changed = self._create_monitores(row, header, dictionaries, equipo, summary, row_data, sheet.title, row_number)
                    asignacion_changed = self._upsert_asignacion(row, header, dictionaries, equipo, funcionario, summary)

                    if funcionario_action == "create" or equipo_action == "create" or monitores_changed == "create" or asignacion_changed == "create":
                        summary.created += 1
                    elif (
                        funcionario_action == "update"
                        or equipo_action == "update"
                        or monitores_changed == "update"
                        or asignacion_changed == "update"
                    ):
                        summary.updated += 1
                    else:
                        summary.skipped += 1
                        self._skip_row(summary, sheet.title, row_number, "Fila sin cambios aplicables.")
            except Exception as exc:
                summary.skipped += 1
                self._skip_row(summary, sheet.title, row_number, str(exc))

    def _upsert_funcionario(self, row, header, summary, row_data):
        rut = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["rut"]))
        nombre = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["nombre"]))
        unidad_desc = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["unidad"]))
        cargo_desc = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["cargo"]))
        tipo_contrato = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["tipo_contrato"]))

        if not rut:
            return None, "skip"

        unidad = None
        if unidad_desc:
            unidad, _ = UnidadFuncionario.objects.get_or_create(descripcion=unidad_desc)

        cargo = None
        if cargo_desc:
            cargo, _ = CargoFuncionario.objects.get_or_create(descripcion=cargo_desc)

        defaults = {
            "nombre_completo": nombre or f"Funcionario {rut}",
            "email_institucional": self._build_placeholder_email(rut),
            "activo": True,
            "codigo_unidad": unidad,
            "codigo_cargo": cargo,
            "tipo_contrato": tipo_contrato or None,
            "raw_excel_data": row_data,
        }

        funcionario, created = Funcionario.objects.get_or_create(rut=rut, defaults=defaults)
        if created:
            summary.funcionarios_created += 1
            self._record_change(funcionario, "create")
            return funcionario, "create"

        changed_fields = []
        if nombre and funcionario.nombre_completo != nombre:
            funcionario.nombre_completo = nombre
            changed_fields.append("nombre_completo")
        if not funcionario.email_institucional:
            funcionario.email_institucional = self._build_placeholder_email(rut)
            changed_fields.append("email_institucional")
        if unidad and funcionario.codigo_unidad_id != unidad.codigo_unidad:
            funcionario.codigo_unidad = unidad
            changed_fields.append("codigo_unidad")
        if cargo and funcionario.codigo_cargo_id != cargo.codigo_cargo:
            funcionario.codigo_cargo = cargo
            changed_fields.append("codigo_cargo")
        if tipo_contrato and funcionario.tipo_contrato != tipo_contrato:
            funcionario.tipo_contrato = tipo_contrato
            changed_fields.append("tipo_contrato")
        if funcionario.activo is not True:
            funcionario.activo = True
            changed_fields.append("activo")
        if funcionario.raw_excel_data != row_data:
            funcionario.raw_excel_data = row_data
            changed_fields.append("raw_excel_data")

        if changed_fields:
            self._record_change(funcionario, "update")
            funcionario.save(update_fields=changed_fields)
            summary.funcionarios_updated += 1
            return funcionario, "update"

        return funcionario, "skip"

    def _upsert_equipo(self, row, header, dictionaries, summary, row_data, sheet_title, row_number):
        numero_serie = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["numero_serie"]))
        numero_inventario = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["numero_inventario"]))

        if not numero_serie and not numero_inventario:
            return None, "skip"

        condicion = self._catalog_from_code(
            CondicionEquipo,
            dictionaries["condicion_equipo_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["condicion_codigo"]),
            "codigo_condicion",
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["condicion_codigo"],
        )
        tipo_equipo = self._catalog_from_code(
            TipoEquipo,
            dictionaries["tipo_equipo_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["tipo_equipo_codigo"]),
            "codigo_tipo",
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["tipo_equipo_codigo"],
        )
        marca = self._catalog_from_code(
            Marca,
            dictionaries["marca_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["marca_codigo"]),
            "codigo_marca",
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["marca_codigo"],
        )
        ram = self._catalog_from_code(
            Ram,
            dictionaries["ram_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["ram_codigo"]),
            "codigo_ram",
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["ram_codigo"],
        )
        procesador = self._catalog_from_code(
            Procesador,
            dictionaries["procesador_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["procesador_codigo"]),
            "codigo_procesador",
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["procesador_codigo"],
        )
        sistema_operativo = self._catalog_from_code(
            SistemaOperativo,
            dictionaries["so_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["so_codigo"]),
            "codigo_so",
            use_nombre=True,
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["so_codigo"],
        )
        tamano_disco = self._catalog_from_code(
            TamanoDiscoCatalogo,
            dictionaries["tamano_disco_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["tamano_disco_codigo"]),
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["tamano_disco_codigo"],
        )
        tipo_disco = self._catalog_from_code(
            TipoDisco,
            dictionaries["tipo_disco_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["tipo_disco_codigo"]),
            "codigo_disco",
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["tipo_disco_codigo"],
        )
        estado_equipo = self._resolve_estado_equipo(
            dictionaries["estado_equipo_dict"],
            self._row_value(row, header, EQUIPOS_COLUMNS["estado_actual"]),
            summary=summary,
            sheet_title=sheet_title,
            row_number=row_number,
            column_name=EQUIPOS_COLUMNS["estado_actual"],
        )
        direccion_oficina_piso = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["direccion_oficina_piso"]))
        ubicacion = self._resolve_ubicacion_from_text(direccion_oficina_piso)

        equipo = None
        if numero_inventario:
            equipo = Equipo.objects.filter(numero_inventario=numero_inventario).first()
        if equipo is None and numero_serie:
            equipo = Equipo.objects.filter(numero_serie=numero_serie).first()

        creating = equipo is None
        if creating:
            equipo = Equipo(
                numero_inventario=numero_inventario or numero_serie,
                numero_serie=numero_serie or numero_inventario,
            )

        changed_fields = []

        for field_name, value in (
            ("numero_inventario", numero_inventario or equipo.numero_inventario),
            ("numero_serie", numero_serie or equipo.numero_serie),
            ("codigo_tipo", tipo_equipo),
            ("codigo_marca", marca),
            ("codigo_ram", ram),
            ("codigo_procesador", procesador),
            ("codigo_so", sistema_operativo),
            ("tamano_disco", tamano_disco),
            ("codigo_disco", tipo_disco),
            ("nombre_equipo", self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["nombre_equipo"])) or equipo.nombre_equipo),
            ("ip_maquina", self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["ip_maquina"])) or equipo.ip_maquina),
            (
                "direccion_oficina_piso",
                direccion_oficina_piso or equipo.direccion_oficina_piso,
            ),
            ("id_ubicacion", ubicacion),
            ("codigo_condicion", condicion),
            ("codigo_estado", estado_equipo),
            ("observaciones", self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["observacion"])) or equipo.observaciones),
            ("raw_excel_data", row_data),
            ("activo", True),
        ):
            current = getattr(equipo, field_name, None)
            if self._field_needs_update(current, value):
                setattr(equipo, field_name, value)
                changed_fields.append(field_name)

        if creating:
            equipo.save()
            self._record_change(equipo, "create")
            summary.equipos_created += 1
            return equipo, "create"
        if changed_fields:
            self._record_change(equipo, "update")
            equipo.save(update_fields=changed_fields)
            summary.equipos_updated += 1
            return equipo, "update"

        return equipo, "skip"

    def _create_monitores(self, row, header, dictionaries, equipo, summary, row_data, sheet_title, row_number):
        result = "skip"
        for idx in (1, 2):
            marca_raw = self._row_value(row, header, EQUIPOS_COLUMNS[f"monitor_{idx}_marca"])
            if self._is_blank(marca_raw):
                continue

            marca_catalogo = self._catalog_from_code(
                MarcaMonitorCatalogo,
                dictionaries["marca_monitor_dict"],
                marca_raw,
                summary=summary,
                sheet_title=sheet_title,
                row_number=row_number,
                column_name=EQUIPOS_COLUMNS[f"monitor_{idx}_marca"],
            )
            pulgada_catalogo = self._catalog_from_code(
                PulgadaMonitorCatalogo,
                dictionaries["pulgada_monitor_dict"],
                self._row_value(row, header, EQUIPOS_COLUMNS[f"monitor_{idx}_pulgada"]),
                summary=summary,
                sheet_title=sheet_title,
                row_number=row_number,
                column_name=EQUIPOS_COLUMNS[f"monitor_{idx}_pulgada"],
            )

            numero_serie = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS[f"monitor_{idx}_serie"]))
            sigac_monitor = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS[f"monitor_{idx}_sigac"]))
            if numero_serie:
                monitor, created = EquipoMonitor.objects.get_or_create(
                    numero_serie_monitor=numero_serie,
                    defaults={"id_equipo": equipo},
                )
            else:
                monitor = EquipoMonitor(id_equipo=equipo)
                created = True

            changed_fields = []
            if monitor.id_equipo_id != equipo.id_equipo:
                monitor.id_equipo = equipo
                changed_fields.append("id_equipo")
            marca_text = marca_catalogo.descripcion if marca_catalogo else self._clean_string(marca_raw)
            if marca_text and monitor.marca_monitor != marca_text:
                monitor.marca_monitor = marca_text
                changed_fields.append("marca_monitor")
            if marca_catalogo and monitor.marca_monitor_catalogo_id != marca_catalogo.id:
                monitor.marca_monitor_catalogo = marca_catalogo
                changed_fields.append("marca_monitor_catalogo")
            if pulgada_catalogo and monitor.pulgada_monitor_catalogo_id != pulgada_catalogo.id:
                monitor.pulgada_monitor_catalogo = pulgada_catalogo
                changed_fields.append("pulgada_monitor_catalogo")
            pulgada_num = self._parse_monitor_inches(pulgada_catalogo.descripcion if pulgada_catalogo else self._row_value(row, header, EQUIPOS_COLUMNS[f"monitor_{idx}_pulgada"]))
            if pulgada_num is not None and monitor.pulgadas != pulgada_num:
                monitor.pulgadas = pulgada_num
                changed_fields.append("pulgadas")
            if sigac_monitor and monitor.sigac_monitor != sigac_monitor:
                monitor.sigac_monitor = sigac_monitor
                changed_fields.append("sigac_monitor")
            if monitor.activo is not True:
                monitor.activo = True
                changed_fields.append("activo")

            if created:
                monitor.save()
                self._record_change(monitor, "create")
                summary.monitores_created += 1
                result = "create"
            elif changed_fields:
                self._record_change(monitor, "update")
                monitor.save(update_fields=changed_fields)
                if result != "create":
                    result = "update"
        return result

    def _upsert_asignacion(self, row, header, dictionaries, equipo, funcionario, summary):
        rut = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["rut"]))
        estado_raw = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["estado_actual"]))
        estado_normalized = (estado_raw or "").strip().upper()

        if not rut:
            return "skip"
        if estado_normalized == "BODEGA":
            return "skip"
        if funcionario is None:
            return "skip"

        active_qs = AsignacionEquipo.objects.filter(
            id_equipo=equipo,
            activo=True,
            fecha_devolucion__isnull=True,
        ).order_by("-id_asignacion")

        active_current = active_qs.filter(id_funcionario=funcionario).first()
        if active_current:
            return "skip"

        fecha_asignacion = self._parse_excel_date(
            self._row_value(row, header, EQUIPOS_COLUMNS["fecha_asignacion_actualizacion"])
        ) or timezone.now().date()
        fecha_entrega = self._parse_excel_date(
            self._row_value(row, header, EQUIPOS_COLUMNS["fecha_entrega"])
        )
        numero_acta = self._clean_string(self._row_value(row, header, EQUIPOS_COLUMNS["numero_acta"]))
        responsable = self._clean_string(
            self._row_value(row, header, EQUIPOS_COLUMNS["nombre_responsable_asignacion"])
        )

        previous = active_qs.first()
        if previous:
            self._record_change(previous, "update")
            previous.fecha_devolucion = fecha_asignacion
            previous.estado_asignacion = "Cerrada"
            previous.activo = False
            previous.save(update_fields=["fecha_devolucion", "estado_asignacion", "activo"])

        asignacion = AsignacionEquipo.objects.create(
            id_equipo=equipo,
            id_funcionario=funcionario,
            fecha_asignacion=fecha_asignacion,
            estado_asignacion="Activa",
            activo=True,
            motivo_asignacion=f"Importado desde {self.path.name}",
            acta_entrega=numero_acta or None,
            nombre_responsable_entrega=responsable or None,
            fecha_entrega_excel=fecha_entrega,
        )
        self._record_change(asignacion, "create")
        summary.asignaciones_created += 1
        return "create"

    def _catalog_from_code(
        self,
        model,
        dictionary,
        raw_value,
        pk_field="id",
        use_nombre=False,
        summary=None,
        sheet_title=None,
        row_number=None,
        column_name=None,
    ):
        code = self._to_int_or_none(raw_value)
        if code is None:
            return None

        descripcion = dictionary.get(code)
        if not descripcion:
            if summary is not None and sheet_title and row_number and column_name:
                summary.errores.append(
                    f'Hoja "{sheet_title}" fila {row_number}: no existe código {code} en CODIGOS para columna "{column_name}".'
                )
            return None

        lookup = {"descripcion": descripcion}
        defaults = {"codigo": code} if hasattr(model, "codigo") else {}

        if model is SistemaOperativo and use_nombre:
            lookup = {"nombre": descripcion}
            defaults["descripcion"] = descripcion

        obj, _ = model.objects.get_or_create(**lookup, defaults=defaults)

        changed_fields = []
        if hasattr(obj, "codigo") and obj.codigo != code:
            obj.codigo = code
            changed_fields.append("codigo")
        if model is SistemaOperativo and use_nombre and obj.descripcion != descripcion:
            obj.descripcion = descripcion
            changed_fields.append("descripcion")

        if changed_fields:
            obj.save(update_fields=changed_fields)

        return obj

    def _resolve_estado_equipo(self, dictionary, raw_value, summary=None, sheet_title=None, row_number=None, column_name=None):
        text = self._clean_string(raw_value)
        code = self._to_int_or_none(raw_value)
        if code is not None:
            return self._catalog_from_code(
                EstadoEquipo,
                dictionary,
                code,
                "codigo_estado",
                summary=summary,
                sheet_title=sheet_title,
                row_number=row_number,
                column_name=column_name,
            )

        if not text:
            text = "ASIGNADO"

        normalized = text.upper()
        existing = EstadoEquipo.objects.filter(descripcion__iexact=normalized).first()
        if existing:
            return existing

        permite_asignacion = normalized != "ASIGNADO"
        return EstadoEquipo.objects.create(
            descripcion=normalized,
            permite_asignacion=permite_asignacion,
        )

    def _header_map(self, sheet):
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        return {
            self._normalize_header(value): idx
            for idx, value in enumerate(first_row)
            if not self._is_blank(value)
        }

    def _original_headers(self, sheet):
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        return [self._clean_string(value) for value in first_row]

    def _row_to_dict(self, row, original_headers, sheet_title, row_number, dictionaries):
        data = {
            "__sheet_name": sheet_title,
            "__excel_row": row_number,
        }
        for idx, header in enumerate(original_headers):
            key = (header or "").strip()
            value = row[idx] if idx < len(row) else None
            if not key or self._is_blank(value):
                continue
            data[key] = self._display_excel_value(key, value, dictionaries)
        return data

    def _row_value(self, row, header, column_name):
        idx = header.get(self._normalize_header(column_name))
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def _skip_row(self, summary, sheet_title, row_number, reason):
        summary.skipped_rows.append(
            {
                "sheet": sheet_title,
                "row": row_number,
                "reason": reason,
            }
        )

    def _parse_code_and_description(self, raw_value):
        text = self._clean_string(raw_value)
        if not text:
            return None, None

        match = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", text)
        if not match:
            return None, text

        return int(match.group(1)), match.group(2).strip()

    def _to_int_or_none(self, raw_value):
        if self._is_blank(raw_value):
            return None
        if isinstance(raw_value, bool):
            return None
        if isinstance(raw_value, (int, float)):
            return int(raw_value)

        text = self._clean_string(raw_value)
        if not text:
            return None

        code, _ = self._parse_code_and_description(text)
        if code is not None:
            return code

        if re.fullmatch(r"\d+", text):
            return int(text)
        return None

    def _parse_monitor_inches(self, raw_value):
        text = self._clean_string(raw_value)
        if not text:
            return None
        match = re.search(r"(\d+)", text)
        return int(match.group(1)) if match else None

    def _parse_excel_date(self, raw_value):
        if self._is_blank(raw_value):
            return None
        if isinstance(raw_value, datetime):
            return raw_value.date()
        if isinstance(raw_value, date):
            return raw_value

        text = self._clean_string(raw_value)
        if not text:
            return None

        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def _build_placeholder_email(self, rut):
        normalized = re.sub(r"[^0-9kK]", "", rut or "").lower()
        if not normalized:
            normalized = "sinrut"
        return f"{normalized}@import.local"

    def _resolve_ubicacion_from_text(self, raw_value):
        text = self._normalize_lookup_text(raw_value)
        if not text:
            return None

        matched_sede = None
        for sede, aliases in SEDE_ALIASES.items():
            normalized_aliases = [self._normalize_lookup_text(alias) for alias in aliases]
            if any(text.endswith(alias) or f" {alias} " in f" {text} " for alias in normalized_aliases):
                matched_sede = sede
                break

        if not matched_sede:
            return None

        ubicacion, _ = Ubicacion.objects.get_or_create(
            nombre_sede=matched_sede,
            defaults={"activo": True},
        )
        if ubicacion.activo is not True:
            ubicacion.activo = True
            ubicacion.save(update_fields=["activo"])
        return ubicacion

    def _field_needs_update(self, current, new_value):
        if new_value is None:
            return False
        current_id = getattr(current, "pk", current)
        new_id = getattr(new_value, "pk", new_value)
        return current_id != new_id

    def _row_is_empty(self, row):
        return all(self._is_blank(value) for value in row)

    def _is_blank(self, value):
        if value is None:
            return True
        if isinstance(value, str):
            text = value.strip()
            return text == "" or text.lower() == "nan"
        return False

    def _clean_string(self, value):
        if self._is_blank(value):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _normalize_cell_value(self, value):
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    def _display_excel_value(self, header, value, dictionaries):
        normalized_header = self._normalize_header(header)
        mapping = {
            self._normalize_header(EQUIPOS_COLUMNS["condicion_codigo"]): "condicion_equipo_dict",
            self._normalize_header(EQUIPOS_COLUMNS["tipo_equipo_codigo"]): "tipo_equipo_dict",
            self._normalize_header(EQUIPOS_COLUMNS["marca_codigo"]): "marca_dict",
            self._normalize_header(EQUIPOS_COLUMNS["ram_codigo"]): "ram_dict",
            self._normalize_header(EQUIPOS_COLUMNS["procesador_codigo"]): "procesador_dict",
            self._normalize_header(EQUIPOS_COLUMNS["so_codigo"]): "so_dict",
            self._normalize_header(EQUIPOS_COLUMNS["tamano_disco_codigo"]): "tamano_disco_dict",
            self._normalize_header(EQUIPOS_COLUMNS["tipo_disco_codigo"]): "tipo_disco_dict",
            self._normalize_header(EQUIPOS_COLUMNS["estado_actual"]): "estado_equipo_dict",
            self._normalize_header(EQUIPOS_COLUMNS["monitor_1_marca"]): "marca_monitor_dict",
            self._normalize_header(EQUIPOS_COLUMNS["monitor_2_marca"]): "marca_monitor_dict",
            self._normalize_header(EQUIPOS_COLUMNS["monitor_1_pulgada"]): "pulgada_monitor_dict",
            self._normalize_header(EQUIPOS_COLUMNS["monitor_2_pulgada"]): "pulgada_monitor_dict",
        }
        dict_name = mapping.get(normalized_header)
        if not dict_name:
            return self._normalize_cell_value(value)

        code = self._to_int_or_none(value)
        if code is None:
            return self._normalize_cell_value(value)

        descripcion = dictionaries.get(dict_name, {}).get(code)
        if not descripcion:
            return self._normalize_cell_value(value)
        return descripcion

    def _normalize_header(self, value):
        text = self._clean_string(value).upper()
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def _normalize_lookup_text(self, value):
        text = self._clean_string(value).upper()
        replacements = str.maketrans({
            "Á": "A",
            "É": "E",
            "Í": "I",
            "Ó": "O",
            "Ú": "U",
            "Ü": "U",
        })
        text = text.translate(replacements)
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def _record_change(self, instance, action):
        if self.lote is None:
            return

        snapshot = None
        if action == "update":
            snapshot = self._snapshot_instance(instance)

        ImportacionInventarioRegistro.objects.create(
            id_lote=self.lote,
            modelo=instance.__class__.__name__,
            object_id=instance.pk,
            accion=action,
            snapshot_previo=snapshot,
        )

    def _snapshot_instance(self, instance):
        data = {}
        for field in instance._meta.concrete_fields:
            if field.primary_key:
                continue
            name = field.attname
            data[name] = getattr(instance, name)
        return data


def rollback_import_lote(lote=None):
    target_lote = lote or ImportacionInventarioLote.objects.order_by("-id_lote").first()
    if target_lote is None:
        raise ValueError("No hay lotes de importación registrados para revertir.")

    if target_lote.estado == "revertido":
        raise ValueError(f"El lote #{target_lote.id_lote} ya fue revertido.")

    with transaction.atomic():
        registros = list(target_lote.registros.order_by("-id_registro"))
        for registro in registros:
            model = ExcelInventoryImporter.TRACKED_MODELS.get(registro.modelo)
            if model is None:
                continue

            instance = model.objects.filter(pk=registro.object_id).first()

            if registro.accion == "create":
                if instance is not None:
                    instance.delete()
                continue

            if registro.accion == "update" and instance is not None and registro.snapshot_previo:
                changed_fields = []
                for key, value in registro.snapshot_previo.items():
                    if getattr(instance, key) != value:
                        setattr(instance, key, value)
                        changed_fields.append(key)
                if changed_fields:
                    instance.save(update_fields=changed_fields)

        target_lote.estado = "revertido"
        target_lote.save(update_fields=["estado"])

    return {
        "id_lote": target_lote.id_lote,
        "archivo_origen": target_lote.archivo_origen,
        "registros_revertidos": len(registros),
    }
