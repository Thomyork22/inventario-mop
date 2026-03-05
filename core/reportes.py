from io import BytesIO
import csv
import os
from datetime import datetime, timedelta
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .permissions import IsInMopInventarioGroup
from .models import Equipo, AsignacionEquipo, Mantenimiento, EquipoMonitor


def _equipo_modelo_display(equipo):
    if not equipo:
        return ""
    raw = getattr(equipo, "raw_excel_data", {}) or {}
    marca = (getattr(getattr(equipo, "codigo_marca", None), "descripcion", None) or "").strip().upper()

    modelo = (getattr(equipo, "modelo", None) or "").strip()
    nombre = (getattr(equipo, "nombre_equipo", None) or "").strip()
    raw_nombre = (
        str(raw.get("NOMBRE DE EQUIPO") or "").strip()
        or str(raw.get("NOMBRE EQUIPO") or "").strip()
        or str(raw.get("MODELO") or "").strip()
    )

    def is_brand_only(value):
        return bool(value and marca and value.strip().upper() == marca)

    if modelo and not is_brand_only(modelo):
        return modelo
    if nombre and not is_brand_only(nombre):
        return nombre
    if raw_nombre and not is_brand_only(raw_nombre):
        return raw_nombre
    if modelo:
        return modelo
    if nombre:
        return nombre
    if raw_nombre:
        return raw_nombre
    if marca:
        return marca.title()
    return "Sin modelo"


def _bool_param(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("1", "true", "t", "si", "sí", "yes"):
        return True
    if s in ("0", "false", "f", "no"):
        return False
    return None


def _date_param(v):
    if not v:
        return None
    try:
        return datetime.strptime(v, "%Y-%m-%d").date()
    except Exception:
        return None


def _csv_response(filename: str, headers: list[str], rows: list[list]):
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(resp)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    return resp


def _xlsx_response(filename: str, sheet_name: str, headers: list[str], rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.append(headers)
    for r in rows:
        ws.append(list(r))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _pdf_response(filename: str, title: str, headers: list[str], rows: list[list], wide=False):
    bio = BytesIO()
    page_size = landscape(A4) if wide else A4
    doc = SimpleDocTemplate(
        bio,
        pagesize=page_size,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Heading3"]
    normal_style = styles["BodyText"]
    normal_style.fontSize = 8
    normal_style.leading = 10

    story = [
        Paragraph(title, title_style),
        Spacer(1, 5 * mm),
    ]

    table_data = [headers]
    for row in rows:
        normalized_row = []
        for value in row:
            text = "" if value is None else str(value)
            normalized_row.append(Paragraph(text.replace("\n", "<br/>"), normal_style))
        table_data.append(normalized_row)

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    story.append(table)
    doc.build(story)
    bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# =========================
# Inventario
# =========================

class ReporteInventarioCSV(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = Equipo.objects.all().order_by("id_equipo")

        activo = _bool_param(request.query_params.get("activo"))
        if activo is not None:
            qs = qs.filter(activo=activo)

        id_ubicacion = request.query_params.get("id_ubicacion")
        if id_ubicacion:
            qs = qs.filter(id_ubicacion_id=id_ubicacion)

        codigo_estado = request.query_params.get("codigo_estado")
        if codigo_estado:
            qs = qs.filter(codigo_estado_id=codigo_estado)

        headers = [
            "id_equipo", "numero_inventario", "numero_serie", "modelo",
            "estado", "sede", "activo", "fecha_compra"
        ]

        rows = []
        for e in qs:
            rows.append([
                e.id_equipo,
                e.numero_inventario,
                e.numero_serie,
                _equipo_modelo_display(e),
                getattr(e.codigo_estado, "descripcion", "") if e.codigo_estado else "",
                getattr(e.id_ubicacion, "nombre_sede", "") if e.id_ubicacion else "",
                "SI" if e.activo else "NO",
                e.fecha_compra.isoformat() if e.fecha_compra else "",
            ])

        return _csv_response("reporte_inventario.csv", headers, rows)


class ReporteInventarioXLSX(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = Equipo.objects.all().order_by("id_equipo")

        activo = _bool_param(request.query_params.get("activo"))
        if activo is not None:
            qs = qs.filter(activo=activo)

        id_ubicacion = request.query_params.get("id_ubicacion")
        if id_ubicacion:
            qs = qs.filter(id_ubicacion_id=id_ubicacion)

        codigo_estado = request.query_params.get("codigo_estado")
        if codigo_estado:
            qs = qs.filter(codigo_estado_id=codigo_estado)

        headers = [
            "id_equipo", "numero_inventario", "numero_serie", "modelo",
            "estado", "sede", "activo", "fecha_compra"
        ]

        rows = []
        for e in qs:
            rows.append([
                e.id_equipo,
                e.numero_inventario,
                e.numero_serie,
                _equipo_modelo_display(e),
                getattr(e.codigo_estado, "descripcion", "") if e.codigo_estado else "",
                getattr(e.id_ubicacion, "nombre_sede", "") if e.id_ubicacion else "",
                "SI" if e.activo else "NO",
                e.fecha_compra.isoformat() if e.fecha_compra else "",
            ])

        return _xlsx_response("reporte_inventario.xlsx", "Inventario", headers, rows)


class ReporteInventarioPDF(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = Equipo.objects.all().order_by("id_equipo")

        activo = _bool_param(request.query_params.get("activo"))
        if activo is not None:
            qs = qs.filter(activo=activo)

        id_ubicacion = request.query_params.get("id_ubicacion")
        if id_ubicacion:
            qs = qs.filter(id_ubicacion_id=id_ubicacion)

        codigo_estado = request.query_params.get("codigo_estado")
        if codigo_estado:
            qs = qs.filter(codigo_estado_id=codigo_estado)

        headers = [
            "ID", "Inventario", "Serie", "Modelo", "Estado", "Sede", "Activo", "Fecha compra"
        ]
        rows = []
        for e in qs:
            rows.append([
                e.id_equipo,
                e.numero_inventario,
                e.numero_serie,
                _equipo_modelo_display(e),
                getattr(e.codigo_estado, "descripcion", "") if e.codigo_estado else "",
                getattr(e.id_ubicacion, "nombre_sede", "") if e.id_ubicacion else "",
                "SI" if e.activo else "NO",
                e.fecha_compra.isoformat() if e.fecha_compra else "",
            ])
        return _pdf_response("reporte_inventario.pdf", "Reporte Inventario", headers, rows, wide=True)


# =========================
# Asignaciones
# =========================

class ReporteAsignacionesCSV(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = AsignacionEquipo.objects.select_related("id_equipo", "id_funcionario").order_by("-id_asignacion")

        solo_activas = _bool_param(request.query_params.get("solo_activas"))
        if solo_activas is True:
            qs = qs.filter(activo=True, fecha_devolucion__isnull=True)

        headers = ["id_asignacion", "funcionario", "rut", "equipo", "modelo", "fecha_asignacion", "fecha_devolucion", "activo"]
        rows = []
        for a in qs:
            f = a.id_funcionario
            e = a.id_equipo
            rows.append([
                a.id_asignacion,
                getattr(f, "nombre_completo", "") if f else "",
                getattr(f, "rut", "") if f else "",
                getattr(e, "numero_inventario", "") if e else "",
                getattr(e, "modelo", "") if e else "",
                a.fecha_asignacion.isoformat() if a.fecha_asignacion else "",
                a.fecha_devolucion.isoformat() if a.fecha_devolucion else "",
                "SI" if a.activo else "NO",
            ])
        return _csv_response("reporte_asignaciones.csv", headers, rows)


class ReporteAsignacionesXLSX(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = AsignacionEquipo.objects.select_related("id_equipo", "id_funcionario").order_by("-id_asignacion")

        solo_activas = _bool_param(request.query_params.get("solo_activas"))
        if solo_activas is True:
            qs = qs.filter(activo=True, fecha_devolucion__isnull=True)

        template_path = _resolve_asignaciones_template_path()
        if template_path:
            wb = load_workbook(template_path)
            ws = wb["EQUIPOS ACTIVOS"] if "EQUIPOS ACTIVOS" in wb.sheetnames else wb.active

            max_row_to_clean = max(ws.max_row, 2500)
            for row_idx in range(2, max_row_to_clean + 1):
                for col_idx in range(1, 28):  # A..AA
                    ws.cell(row=row_idx, column=col_idx).value = None

            for idx, a in enumerate(qs, start=2):
                f = a.id_funcionario
                e = a.id_equipo
                raw = getattr(e, "raw_excel_data", {}) or {}
                monitores = (
                    list(EquipoMonitor.objects.filter(id_equipo=e).order_by("id_equipo_monitor")[:2])
                    if e
                    else []
                )
                m1 = monitores[0] if len(monitores) >= 1 else None
                m2 = monitores[1] if len(monitores) >= 2 else None

                direccion = (getattr(e, "direccion_oficina_piso", "") if e else "") or str(
                    raw.get("DIRECCIÓN OFICINA / PISO") or ""
                ).strip()
                region = (
                    getattr(getattr(getattr(e, "id_ubicacion", None), "codigo_region", None), "nombre", "")
                    if e
                    else ""
                ) or str(raw.get("REGIÓN") or "")

                row_values = {
                    1: region,
                    2: getattr(getattr(e, "codigo_tipo", None), "descripcion", "") if e else "",
                    3: getattr(getattr(e, "codigo_marca", None), "descripcion", "") if e else "",
                    4: getattr(getattr(e, "codigo_ram", None), "descripcion", "") if e else "",
                    5: getattr(getattr(e, "codigo_procesador", None), "descripcion", "") if e else "",
                    6: getattr(getattr(e, "codigo_so", None), "descripcion", "") if e else "",
                    7: getattr(getattr(e, "tamano_disco", None), "descripcion", "") if e else "",
                    8: getattr(getattr(e, "codigo_disco", None), "descripcion", "") if e else "",
                    9: getattr(e, "numero_serie", "") if e else "",
                    10: getattr(e, "numero_inventario", "") if e else "",
                    11: _equipo_modelo_display(e),
                    12: getattr(e, "ip_maquina", "") if e else "",
                    13: (
                        getattr(getattr(m1, "marca_monitor_catalogo", None), "descripcion", "")
                        or getattr(m1, "marca_monitor", "")
                        or str(raw.get("MARCA MONITOR 1") or "")
                    ),
                    14: getattr(m1, "numero_serie_monitor", "") if m1 else str(raw.get("SERIE MONITOR 1") or ""),
                    15: getattr(m1, "sigac_monitor", "") if m1 else str(raw.get("SIGAC MONITOR 1") or ""),
                    16: (
                        getattr(getattr(m1, "pulgada_monitor_catalogo", None), "descripcion", "")
                        if m1
                        else str(raw.get("PULGADA MONITOR 1") or "")
                    ),
                    17: (
                        getattr(getattr(m2, "marca_monitor_catalogo", None), "descripcion", "")
                        or getattr(m2, "marca_monitor", "")
                        or str(raw.get("MARCA MONITOR 2") or "")
                    ),
                    18: getattr(m2, "numero_serie_monitor", "") if m2 else str(raw.get("SERIE MONITOR 2") or ""),
                    19: getattr(m2, "sigac_monitor", "") if m2 else str(raw.get("SIGAC MONITOR 2") or ""),
                    20: (
                        getattr(getattr(m2, "pulgada_monitor_catalogo", None), "descripcion", "")
                        if m2
                        else str(raw.get("PULGADA MONITOR 2") or "")
                    ),
                    21: a.fecha_asignacion.isoformat() if a.fecha_asignacion else "",
                    22: (a.nombre_responsable_entrega or "").strip() or str(raw.get("NOMBRE RESPONSABLE DE ASIGNACIÓN") or ""),
                    23: getattr(f, "rut", "") if f else "",
                    24: getattr(f, "nombre_completo", "") if f else "",
                    25: direccion,
                    26: str(raw.get("DEPTO / UNIDAD") or ""),
                    27: a.fecha_entrega_excel.isoformat() if a.fecha_entrega_excel else (a.fecha_devolucion.isoformat() if a.fecha_devolucion else ""),
                }
                for col_idx, value in row_values.items():
                    ws.cell(row=idx, column=col_idx).value = value

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(
                bio.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = 'attachment; filename="reporte_asignaciones_formato.xlsx"'
            return resp

        headers = [
            "NOMBRE FUNCIONARIO",
            "RUT",
            "EMAIL",
            "INVENTARIO",
            "SERIE",
            "TIPO EQUIPO",
            "MARCA",
            "DIRECCIÓN OFICINA / PISO",
            "OFICINA",
            "SEDE",
            "DEPTO / UNIDAD",
            "CARGO FUNCIONAL",
            "FECHA ASIGNACIÓN",
            "FECHA DEVOLUCIÓN",
            "ESTADO ASIGNACIÓN",
            "ACTIVO",
        ]
        rows = []
        for a in qs:
            f = a.id_funcionario
            e = a.id_equipo
            raw = getattr(e, "raw_excel_data", {}) or {}
            tipo = getattr(getattr(e, "codigo_tipo", None), "descripcion", "") if e else ""
            marca = getattr(getattr(e, "codigo_marca", None), "descripcion", "") if e else ""
            direccion = getattr(e, "direccion_oficina_piso", "") if e else ""
            if not direccion:
                direccion = str(raw.get("DIRECCIÓN OFICINA / PISO") or "").strip()
            oficina = str(raw.get("OFICINA") or raw.get("PISO OFICINA") or "").strip()
            sede = getattr(getattr(e, "id_ubicacion", None), "nombre_sede", "") if e else ""
            if not sede:
                sede = _infer_sede(direccion)
            unidad = str(raw.get("DEPTO / UNIDAD") or "").strip()
            cargo = str(raw.get("CARGO FUNCIONAL") or "").strip()
            rows.append([
                getattr(f, "nombre_completo", "") if f else "",
                getattr(f, "rut", "") if f else "",
                getattr(f, "email_institucional", "") if f else "",
                getattr(e, "numero_inventario", "") if e else "",
                getattr(e, "numero_serie", "") if e else "",
                tipo,
                marca,
                direccion,
                oficina,
                sede,
                unidad,
                cargo,
                a.fecha_asignacion.isoformat() if a.fecha_asignacion else "",
                a.fecha_devolucion.isoformat() if a.fecha_devolucion else "",
                (a.estado_asignacion or "").strip() or ("Activa" if a.activo else "Cerrada"),
                "SI" if a.activo else "NO",
            ])
        return _xlsx_response("reporte_asignaciones_formato.xlsx", "Asignaciones", headers, rows)


class ReporteAsignacionesPDF(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = AsignacionEquipo.objects.select_related("id_equipo", "id_funcionario").order_by("-id_asignacion")
        solo_activas = _bool_param(request.query_params.get("solo_activas"))
        if solo_activas is True:
            qs = qs.filter(activo=True, fecha_devolucion__isnull=True)

        headers = [
            "Funcionario",
            "RUT",
            "Inventario",
            "Serie",
            "Direccion / Piso",
            "Sede",
            "Unidad",
            "Cargo",
            "Fecha asignacion",
            "Fecha devolucion",
            "Estado",
        ]
        rows = []
        for a in qs:
            f = a.id_funcionario
            e = a.id_equipo
            raw = getattr(e, "raw_excel_data", {}) or {}
            direccion = getattr(e, "direccion_oficina_piso", "") if e else ""
            if not direccion:
                direccion = str(raw.get("DIRECCIÓN OFICINA / PISO") or "").strip()
            sede = getattr(getattr(e, "id_ubicacion", None), "nombre_sede", "") if e else ""
            if not sede:
                sede = _infer_sede(direccion)

            rows.append([
                getattr(f, "nombre_completo", "") if f else "",
                getattr(f, "rut", "") if f else "",
                getattr(e, "numero_inventario", "") if e else "",
                getattr(e, "numero_serie", "") if e else "",
                direccion,
                sede,
                str(raw.get("DEPTO / UNIDAD") or "").strip(),
                str(raw.get("CARGO FUNCIONAL") or "").strip(),
                a.fecha_asignacion.isoformat() if a.fecha_asignacion else "",
                a.fecha_devolucion.isoformat() if a.fecha_devolucion else "",
                (a.estado_asignacion or "").strip() or ("Activa" if a.activo else "Cerrada"),
            ])
        return _pdf_response("reporte_asignaciones.pdf", "Reporte Asignaciones", headers, rows, wide=True)


def _infer_sede(value):
    text = str(value or "").strip().upper()
    if "ANIMAS" in text or "ÁNIMAS" in text:
        return "Ánimas"
    if "YUNGAY" in text:
        return "Yungay"
    if "LA UNION" in text or "LA UNIÓN" in text:
        return "La Unión"
    if "BOUCHEFF" in text or "BOUCHEF" in text:
        return "Bouchéff"
    return ""


def _resolve_asignaciones_template_path():
    candidates = []
    env_path = os.environ.get("ASIGNACIONES_TEMPLATE_PATH")
    if env_path:
        candidates.append(Path(env_path))

    candidates.extend(
        [
            Path(settings.BASE_DIR) / "plantillas" / "Copia de FORMATO_INVENTARIO_2026.xlsx",
            Path(settings.BASE_DIR) / "Copia de FORMATO_INVENTARIO_2026.xlsx",
            Path("/Users/thomyrk/Desktop/mop/Copia de FORMATO_INVENTARIO_2026.xlsx"),
        ]
    )
    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


# =========================
# Mantenimientos
# =========================

class ReporteMantenimientosCSV(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = Mantenimiento.objects.select_related("id_equipo", "codigo_estado_mantenimiento").order_by("-id_mantenimiento")

        headers = ["id_mantenimiento", "equipo", "modelo", "estado", "fecha_solicitud", "fecha_ingreso_taller", "fecha_salida_taller", "fecha_entrega"]
        rows = []
        for m in qs:
            e = m.id_equipo
            est = m.codigo_estado_mantenimiento
            rows.append([
                m.id_mantenimiento,
                getattr(e, "numero_inventario", "") if e else "",
                _equipo_modelo_display(e),
                getattr(est, "descripcion", "") if est else "",
                m.fecha_solicitud.isoformat() if m.fecha_solicitud else "",
                m.fecha_ingreso_taller.isoformat() if m.fecha_ingreso_taller else "",
                m.fecha_salida_taller.isoformat() if m.fecha_salida_taller else "",
                m.fecha_entrega.isoformat() if m.fecha_entrega else "",
            ])
        return _csv_response("reporte_mantenimientos.csv", headers, rows)


class ReporteMantenimientosXLSX(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = Mantenimiento.objects.select_related("id_equipo", "codigo_estado_mantenimiento").order_by("-id_mantenimiento")

        headers = ["id_mantenimiento", "equipo", "modelo", "estado", "fecha_solicitud", "fecha_ingreso_taller", "fecha_salida_taller", "fecha_entrega"]
        rows = []
        for m in qs:
            e = m.id_equipo
            est = m.codigo_estado_mantenimiento
            rows.append([
                m.id_mantenimiento,
                getattr(e, "numero_inventario", "") if e else "",
                _equipo_modelo_display(e),
                getattr(est, "descripcion", "") if est else "",
                m.fecha_solicitud.isoformat() if m.fecha_solicitud else "",
                m.fecha_ingreso_taller.isoformat() if m.fecha_ingreso_taller else "",
                m.fecha_salida_taller.isoformat() if m.fecha_salida_taller else "",
                m.fecha_entrega.isoformat() if m.fecha_entrega else "",
            ])
        return _xlsx_response("reporte_mantenimientos.xlsx", "Mantenimientos", headers, rows)


class ReporteMantenimientosPDF(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        qs = Mantenimiento.objects.select_related("id_equipo", "codigo_estado_mantenimiento").order_by("-id_mantenimiento")
        headers = [
            "ID", "Inventario", "Modelo", "Estado", "Fecha solicitud", "Ingreso taller", "Salida taller", "Entrega"
        ]
        rows = []
        for m in qs:
            e = m.id_equipo
            est = m.codigo_estado_mantenimiento
            rows.append([
                m.id_mantenimiento,
                getattr(e, "numero_inventario", "") if e else "",
                _equipo_modelo_display(e),
                getattr(est, "descripcion", "") if est else "",
                m.fecha_solicitud.isoformat() if m.fecha_solicitud else "",
                m.fecha_ingreso_taller.isoformat() if m.fecha_ingreso_taller else "",
                m.fecha_salida_taller.isoformat() if m.fecha_salida_taller else "",
                m.fecha_entrega.isoformat() if m.fecha_entrega else "",
            ])
        return _pdf_response("reporte_mantenimientos.pdf", "Reporte Mantenimientos", headers, rows, wide=True)


# =========================
# Garantías
# =========================

class ReporteGarantiasCSV(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        hoy = timezone.now().date()
        dias = request.query_params.get("dias")  # default 30
        try:
            dias = int(dias) if dias else 30
        except Exception:
            dias = 30

        tipo = (request.query_params.get("tipo") or "pv").lower()  # pv | vencidas
        limite = hoy + timedelta(days=dias)

        qs = Equipo.objects.filter(fecha_fin_garantia__isnull=False).order_by("fecha_fin_garantia")

        if tipo == "vencidas":
            qs = qs.filter(fecha_fin_garantia__lt=hoy)
        else:
            qs = qs.filter(fecha_fin_garantia__range=(hoy, limite))

        headers = ["id_equipo", "numero_inventario", "modelo", "fecha_fin_garantia", "dias_restantes"]
        rows = []
        for e in qs:
            dias_rest = (e.fecha_fin_garantia - hoy).days if e.fecha_fin_garantia else ""
            rows.append([e.id_equipo, e.numero_inventario, _equipo_modelo_display(e), e.fecha_fin_garantia.isoformat(), dias_rest])
        return _csv_response("reporte_garantias.csv", headers, rows)


class ReporteGarantiasXLSX(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        hoy = timezone.now().date()
        dias = request.query_params.get("dias")
        try:
            dias = int(dias) if dias else 30
        except Exception:
            dias = 30

        tipo = (request.query_params.get("tipo") or "pv").lower()
        limite = hoy + timedelta(days=dias)

        qs = Equipo.objects.filter(fecha_fin_garantia__isnull=False).order_by("fecha_fin_garantia")
        if tipo == "vencidas":
            qs = qs.filter(fecha_fin_garantia__lt=hoy)
        else:
            qs = qs.filter(fecha_fin_garantia__range=(hoy, limite))

        headers = ["id_equipo", "numero_inventario", "modelo", "fecha_fin_garantia", "dias_restantes"]
        rows = []
        for e in qs:
            dias_rest = (e.fecha_fin_garantia - hoy).days if e.fecha_fin_garantia else ""
            rows.append([e.id_equipo, e.numero_inventario, _equipo_modelo_display(e), e.fecha_fin_garantia.isoformat(), dias_rest])

        return _xlsx_response("reporte_garantias.xlsx", "Garantias", headers, rows)


class ReporteGarantiasPDF(APIView):
    permission_classes = [IsAuthenticated, IsInMopInventarioGroup]

    def get(self, request):
        hoy = timezone.now().date()
        dias = request.query_params.get("dias")
        try:
            dias = int(dias) if dias else 30
        except Exception:
            dias = 30

        tipo = (request.query_params.get("tipo") or "pv").lower()
        limite = hoy + timedelta(days=dias)

        qs = Equipo.objects.filter(fecha_fin_garantia__isnull=False).order_by("fecha_fin_garantia")
        if tipo == "vencidas":
            qs = qs.filter(fecha_fin_garantia__lt=hoy)
        else:
            qs = qs.filter(fecha_fin_garantia__range=(hoy, limite))

        headers = ["ID", "Inventario", "Modelo", "Fecha fin garantia", "Dias restantes"]
        rows = []
        for e in qs:
            dias_rest = (e.fecha_fin_garantia - hoy).days if e.fecha_fin_garantia else ""
            rows.append([e.id_equipo, e.numero_inventario, _equipo_modelo_display(e), e.fecha_fin_garantia.isoformat(), dias_rest])
        return _pdf_response("reporte_garantias.pdf", "Reporte Garantias", headers, rows, wide=False)
