from __future__ import annotations

import re
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction

from openpyxl import load_workbook

from core.models import (
    CondicionEquipo,
    EstadoEquipo,
    Marca,
    MarcaMonitorCatalogo,
    Procesador,
    PulgadaMonitorCatalogo,
    Ram,
    SistemaOperativo,
    TamanoDiscoCatalogo,
    TipoDisco,
    TipoEquipo,
    Ubicacion,
)


FIXED_SEDES = ("Ánimas", "Yungay", "La Unión", "Bouchéff")

FALLBACK_DATA = {
    "condiciones_equipo": ["1 - Bueno", "2 - Regular", "0 - N/A"],
    "estados_equipo": ["1 - Libre", "2 - Ocupado", "3 - En mantenimiento", "4 - Pendiente"],
    "tipos_equipo": ["1 - Notebook", "2 - Desktop", "3 - Monitor", "0 - N/A"],
    "marcas": ["1 - Dell", "2 - HP", "3 - Lenovo", "0 - N/A"],
    "ram": ["1 - 4 GB", "2 - 8 GB", "3 - 16 GB", "0 - N/A"],
    "procesadores": ["1 - Intel Core i5", "2 - Intel Core i7", "3 - AMD Ryzen 5", "0 - N/A"],
    "sistemas_operativos": ["1 - Windows 10", "2 - Windows 11", "3 - Linux", "0 - N/A"],
    "tamanos_disco": ["1 - 256 GB", "2 - 512 GB", "3 - 1 TB", "0 - N/A"],
    "tipos_disco": ["1 - HDD", "2 - SSD", "0 - N/A"],
    "marcas_monitor": ["1 - LG", "2 - Samsung", "3 - Dell", "0 - N/A"],
    "pulgadas_monitor": ["1 - 22", "2 - 24", "3 - 27", "0 - N/A"],
}

COLUMN_ALIASES = {
    "condiciones_equipo": (
        "condición equipo 1 / 2",
        "condicion equipo 1 / 2",
        "condición equipo 1/2",
        "condicion equipo 1/2",
    ),
    "estados_equipo": (
        "estado/equipo (pendiente)",
        "estado/equipo",
        "estado equipo",
    ),
    "tipos_equipo": (
        "tipo equipo 1/2/otro",
        "tipo equipo 1 / 2 / otro",
        "tipo equipo",
    ),
    "marcas": ("marca",),
    "ram": ("ram",),
    "procesadores": ("procesador",),
    "sistemas_operativos": ("s.o", "so", "s/o", "sistema operativo"),
    "tamanos_disco": ("tamaño disco", "tamano disco"),
    "tipos_disco": ("tipo disco 1/2", "tipo disco 1 / 2", "tipo disco"),
    "marcas_monitor": ("marca monitor",),
    "pulgadas_monitor": ("pulgada monitor", "pulgadas monitor"),
}


class Command(BaseCommand):
    help = "Carga catálogos desde la hoja CODIGOS del Excel de inventario."

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            dest="path",
            default="Copia de INVENTARIO REGION XIV 2025.xlsx",
            help='Ruta al archivo Excel. Por defecto: "Copia de INVENTARIO REGION XIV 2025.xlsx".',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        excel_path = Path(options["path"]).expanduser()
        if not excel_path.is_absolute():
            excel_path = Path.cwd() / excel_path

        extracted = self._load_from_excel(excel_path)
        if extracted is None:
            self.stdout.write(
                self.style.WARNING(
                    f'No se encontró un Excel válido en "{excel_path}". Se cargará fallback mínimo.'
                )
            )
            extracted = {key: list(values) for key, values in FALLBACK_DATA.items()}

        summary = {
            "condiciones_equipo": self._seed_condiciones(extracted["condiciones_equipo"]),
            "estados_equipo": self._seed_estados(extracted["estados_equipo"]),
            "tipos_equipo": self._seed_simple(TipoEquipo, extracted["tipos_equipo"]),
            "marcas": self._seed_simple(Marca, extracted["marcas"]),
            "ram": self._seed_ram(extracted["ram"]),
            "procesadores": self._seed_simple(Procesador, extracted["procesadores"]),
            "sistemas_operativos": self._seed_sistemas_operativos(extracted["sistemas_operativos"]),
            "tamanos_disco": self._seed_simple(TamanoDiscoCatalogo, extracted["tamanos_disco"]),
            "tipos_disco": self._seed_simple(TipoDisco, extracted["tipos_disco"]),
            "marcas_monitor": self._seed_simple(MarcaMonitorCatalogo, extracted["marcas_monitor"]),
            "pulgadas_monitor": self._seed_simple(PulgadaMonitorCatalogo, extracted["pulgadas_monitor"]),
            "sedes": self._seed_sedes(),
        }

        self.stdout.write(self.style.SUCCESS("Carga de catálogos completada."))
        for key, value in summary.items():
            self.stdout.write(f"- {key}: {value}")

    def _load_from_excel(self, path: Path):
        if not path.exists():
            return None

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return None

        if "CODIGOS" not in wb.sheetnames:
            return None

        ws = wb["CODIGOS"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None

        header_row_idx, indexes = self._find_header_row(rows)
        if header_row_idx is None:
            return None

        extracted = {key: set() for key in COLUMN_ALIASES}
        for row in rows[header_row_idx + 1:]:
            for key, idx in indexes.items():
                if idx >= len(row):
                    continue
                raw_value = row[idx]
                if raw_value is None:
                    continue
                value = str(raw_value).strip()
                if value:
                    extracted[key].add(value)

        normalized = {
            key: sorted(values, key=lambda item: self._sort_key(item))
            for key, values in extracted.items()
        }
        return {key: normalized.get(key) or list(FALLBACK_DATA[key]) for key in COLUMN_ALIASES}

    def _find_header_row(self, rows):
        for row_idx, row in enumerate(rows[:10]):
            normalized_row = [self._normalize_header(value) for value in row]
            indexes = {}
            for key, aliases in COLUMN_ALIASES.items():
                for idx, candidate in enumerate(normalized_row):
                    if candidate in aliases:
                        indexes[key] = idx
                        break
            if indexes:
                for key in COLUMN_ALIASES:
                    indexes.setdefault(key, None)
                resolved = {key: idx for key, idx in indexes.items() if idx is not None}
                return row_idx, resolved
        return None, {}

    def _seed_simple(self, model, values):
        created = 0
        updated = 0

        for raw_value in values:
            codigo, descripcion = self._parse_code_and_description(raw_value)
            defaults = {}
            if hasattr(model, "codigo"):
                defaults["codigo"] = codigo

            obj, was_created = model.objects.get_or_create(
                descripcion=descripcion,
                defaults=defaults,
            )
            if was_created:
                created += 1
                continue

            if hasattr(obj, "codigo") and codigo is not None and obj.codigo != codigo:
                obj.codigo = codigo
                obj.save(update_fields=["codigo"])
                updated += 1

        return f"{created} creados, {updated} actualizados, {model.objects.count()} total"

    def _seed_condiciones(self, values):
        created = 0
        updated = 0

        for raw_value in values:
            codigo, descripcion = self._parse_code_and_description(raw_value)
            obj, was_created = CondicionEquipo.objects.get_or_create(
                descripcion=descripcion,
                defaults={"codigo": codigo},
            )
            if was_created:
                created += 1
                continue

            changed = False
            if codigo is not None and obj.codigo != codigo:
                obj.codigo = codigo
                changed = True
            if changed:
                obj.save(update_fields=["codigo"])
                updated += 1

        return f"{created} creados, {updated} actualizados, {CondicionEquipo.objects.count()} total"

    def _seed_estados(self, values):
        created = 0
        updated = 0

        for raw_value in values:
            codigo, descripcion = self._parse_code_and_description(raw_value)
            permite_asignacion = self._permite_asignacion(descripcion)
            obj, was_created = EstadoEquipo.objects.get_or_create(
                descripcion=descripcion,
                defaults={
                    "codigo": codigo,
                    "permite_asignacion": permite_asignacion,
                },
            )
            if was_created:
                created += 1
                continue

            changed = False
            if codigo is not None and obj.codigo != codigo:
                obj.codigo = codigo
                changed = True
            if obj.permite_asignacion != permite_asignacion:
                obj.permite_asignacion = permite_asignacion
                changed = True
            if changed:
                obj.save(update_fields=["codigo", "permite_asignacion"])
                updated += 1

        return f"{created} creados, {updated} actualizados, {EstadoEquipo.objects.count()} total"

    def _seed_ram(self, values):
        created = 0
        updated = 0

        for raw_value in values:
            codigo, descripcion = self._parse_code_and_description(raw_value)
            obj, was_created = Ram.objects.get_or_create(
                descripcion=descripcion,
                defaults={"codigo": codigo},
            )
            if was_created:
                created += 1
                continue

            if codigo is not None and obj.codigo != codigo:
                obj.codigo = codigo
                obj.save(update_fields=["codigo"])
                updated += 1

        return f"{created} creados, {updated} actualizados, {Ram.objects.count()} total"

    def _seed_sistemas_operativos(self, values):
        created = 0
        updated = 0

        for raw_value in values:
            codigo, descripcion = self._parse_code_and_description(raw_value)
            obj, was_created = SistemaOperativo.objects.get_or_create(
                nombre=descripcion,
                defaults={"codigo": codigo, "descripcion": descripcion},
            )
            if was_created:
                created += 1
                continue

            changed = False
            if codigo is not None and obj.codigo != codigo:
                obj.codigo = codigo
                changed = True
            if obj.descripcion != descripcion:
                obj.descripcion = descripcion
                changed = True
            if changed:
                obj.save(update_fields=["codigo", "descripcion"])
                updated += 1

        return f"{created} creados, {updated} actualizados, {SistemaOperativo.objects.count()} total"

    def _seed_sedes(self):
        created = 0
        updated = 0

        for sede in FIXED_SEDES:
            obj, was_created = Ubicacion.objects.get_or_create(
                nombre_sede=sede,
                defaults={"activo": True},
            )
            if was_created:
                created += 1
                continue

            if obj.activo is not True:
                obj.activo = True
                obj.save(update_fields=["activo"])
                updated += 1

        return f"{created} creados, {updated} actualizados, {Ubicacion.objects.count()} total"

    def _parse_code_and_description(self, raw_value):
        text = str(raw_value).strip()
        match = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", text)
        if match:
            return int(match.group(1)), match.group(2).strip()
        return None, text

    def _permite_asignacion(self, descripcion):
        value = (descripcion or "").strip().lower()
        if not value:
            return None
        blocked_tokens = ("ocup", "manten", "baja", "pend", "repar", "taller", "fuera")
        return not any(token in value for token in blocked_tokens)

    def _normalize_header(self, value):
        text = str(value or "").strip().lower()
        text = text.replace("á", "a").replace("é", "e").replace("í", "i")
        text = text.replace("ó", "o").replace("ú", "u").replace("°", "")
        text = re.sub(r"\s+", " ", text)
        return text

    def _sort_key(self, raw_value):
        codigo, descripcion = self._parse_code_and_description(raw_value)
        return (codigo is None, codigo if codigo is not None else 999999, descripcion.lower())
